"""端到端功能测试(修正版)"""
import sys
sys.path.insert(0, "/Users/a111/Desktop/qa-tool")
from qa_web import app, XLSX_PATH

XLSX_PATH["path"] = "/Users/a111/Downloads/问题.xlsx"
client = app.test_client()
failed = False

def test(name, ok, detail=""):
    sym = "✓" if ok else "✗"
    print(f"{sym} {name}{('  ' + detail) if detail else ''}")
    if not ok:
        global failed
        failed = True

# 1. 主页
r = client.get("/")
test("GET / 返回 200", r.status_code == 200)
test("GET / 含标题", "qa-tool" in r.get_data(as_text=True))

# 2. info
r = client.get("/api/info")
data = r.get_json()
test("info 200 + loaded", r.status_code == 200 and data.get("loaded"))
test("answer_max=2000", data.get("answer_max") == 2000)

# 3. 列表基础
r = client.get("/api/qa")
data = r.get_json()
test(f"列表 36 条(实际 {data['total']})", data["total"] == 36)
max_existing_seq = max(q["seq"] for q in data["items"] if isinstance(q["seq"], int))
test(f"最大序号是 172(实际 {max_existing_seq})", max_existing_seq == 172)

# 4. 搜索(关键词命中:问题 + 答案都搜)
r = client.get("/api/qa?q=dispatchTask")
data = r.get_json()
# 'dispatchTask' 出现:2 个问题标题 + 一些答案内容,实际 4 条
test(f"搜索 'dispatchTask' 至少 2 条(实际 {data['total']})", data["total"] >= 2)
# 搜一个独特关键词
r = client.get("/api/qa?q=循环依赖")
data = r.get_json()
test(f"搜索 '循环依赖' 应 1 条(实际 {data['total']})", data["total"] == 1)
# 搜不存在的
r = client.get("/api/qa?q=__不存在的关键词__")
data = r.get_json()
test(f"搜索不存在关键词应 0 条(实际 {data['total']})", data["total"] == 0)

# 5. 字符计数行为(code point)
def cp(s): return len(s)
test("'a' * 2000 == 2000", cp('a' * 2000) == 2000)
test("'a' * 2001 == 2001", cp('a' * 2001) == 2001)
mixed = "你好 hello 👋 世界"
test(f"中英emoji空格混合: '{mixed}' = {cp(mixed)} 字符", cp(mixed) == 13)
test("'a b c' = 5 (含 2 空格)", cp('a b c') == 5)
test("空串 = 0", cp('') == 0)

# 6. 添加新 Q&A
new_q = "【测试-临时】请忽略"
new_a = "测试答案内容,用于验证功能"
r = client.post("/api/qa", json={"question": new_q, "answer": new_a})
test(f"添加新 Q&A 200(实际 {r.status_code})", r.status_code == 200)
data = r.get_json()
new_seq = data.get("seq")
test(f"新序号 = 173(实际 {new_seq})", new_seq == 173)
test(f"action=added", data.get("action") == "added")

# 7. 重复检测
r = client.post("/api/qa", json={"question": new_q, "answer": "x"})
test(f"重复应 409(实际 {r.status_code})", r.status_code == 409)
data = r.get_json()
test(f"duplicate=True, existing_seq=173(实际 {data.get('existing_seq')})",
     data.get("duplicate") and data.get("existing_seq") == 173)

# 8. 强制覆盖
r = client.post("/api/qa", json={"question": new_q, "answer": "覆盖后", "force_overwrite": True})
test(f"强制覆盖 200(实际 {r.status_code})", r.status_code == 200)
test(f"action=overwritten", r.get_json().get("action") == "overwritten")

# 9. 验证覆盖真的写进 xlsx
r = client.get("/api/qa")
items = r.get_json()["items"]
overwritten = next((q for q in items if q["seq"] == 173), None)
test(f"覆盖后答案= '覆盖后'(实际 '{overwritten['answer'] if overwritten else 'None'}')",
     overwritten and overwritten["answer"] == "覆盖后")

# 10. 2000 字符限制
r = client.post("/api/qa", json={"question": "超长测试", "answer": "x" * 2001})
test(f"2001 字符应 400(实际 {r.status_code})", r.status_code == 400)
test("错误信息含 '2000'", "2000" in r.get_json().get("error", ""))

# 边界 2000
r = client.post("/api/qa", json={"question": "边界-2000", "answer": "y" * 2000})
test(f"正好 2000 字符应 200(实际 {r.status_code})", r.status_code == 200)
seq_2000 = r.get_json().get("seq")
test(f"2000 边界序号=174(实际 {seq_2000})", seq_2000 == 174)

# 11. 空问题
r = client.post("/api/qa", json={"question": "", "answer": "x"})
test(f"空问题应 400(实际 {r.status_code})", r.status_code == 400)

# 12. 切换文件
r = client.post("/api/xlsx", json={"path": "/Users/a111/Downloads/问题_test_backup.xlsx"})
test(f"切换到备份 200(实际 {r.status_code})", r.status_code == 200)
test("切到不存在文件 404", client.post("/api/xlsx", json={"path": "/nope/x.xlsx"}).status_code == 404)

# 切回主文件
client.post("/api/xlsx", json={"path": "/Users/a111/Downloads/问题.xlsx"})

# 13. 清理(精确删除测试数据)
from openpyxl import load_workbook
wb = load_workbook("/Users/a111/Downloads/问题.xlsx")
ws = wb.active
# 找到测试行(用问题名匹配)
to_delete = []
for i, row in enumerate(ws.iter_rows(values_only=False), start=1):
    if row[1].value in (new_q, "边界-2000"):
        to_delete.append(i)
print(f"  (待清理行号: {to_delete})")
for row_idx in sorted(to_delete, reverse=True):
    ws.delete_rows(row_idx, 1)
wb.save("/Users/a111/Downloads/问题.xlsx")

# 验证
wb2 = load_workbook("/Users/a111/Downloads/问题.xlsx", data_only=True)
final = list(wb2.active.iter_rows(values_only=True))
data_count = len(final) - 1
test(f"清理后剩 36 行(实际 {data_count})", data_count == 36)
test(f"表头正确(实际 {final[0]})", final[0] == ("序号", "问题", "答案"))

print()
print("=" * 50)
print("❌ 有失败" if failed else "✅ 全部通过")
sys.exit(1 if failed else 0)
