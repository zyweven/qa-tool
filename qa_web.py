"""
qa-tool: 把 xlsx 当数据库用的 Q&A 管理页面
启动: python qa_web.py [xlsx 路径]
不传路径时(macOS)弹原生文件选择对话框
"""
import os
import sys
import subprocess
from pathlib import Path

from flask import Flask, request, jsonify, render_template
from openpyxl import Workbook, load_workbook

app = Flask(__name__)

# xlsx 路径(全局,启动时设置,运行中可切换)
XLSX_PATH = {"path": None}

# 答案字符上限
ANSWER_MAX = 2000


# ---------- 工具函数 ----------

def count_codepoints(s) -> int:
    """按 Unicode code point 计数(Python 3 str 天然是 code point)"""
    if s is None:
        return 0
    return len(str(s))


def pick_file_macos() -> str | None:
    """用 macOS 原生对话框选文件(失败返回 None)

    不限制文件类型(现代 macOS 的 of type 需要 UTI,容易选不到),
    改成选所有文件,后端用 validate_xlsx 校验后缀和格式。
    """
    script = '''
    set theFile to choose file with prompt "选择 Q&A xlsx 文件"
    return POSIX path of theFile
    '''
    try:
        result = subprocess.run(
            ["osascript", "-e", script],
            capture_output=True, text=True, timeout=120,
        )
        if result.returncode == 0:
            path = result.stdout.strip()
            if path:
                return path
    except (FileNotFoundError, subprocess.TimeoutExpired):
        pass
    return None


def pick_save_path_macos(default_name: str = "问答.xlsx") -> str | None:
    """用 macOS 原生"保存"对话框,返回用户选中的完整路径"""
    # 转义 AppleScript 字符串里的双引号
    safe_name = default_name.replace('"', '\\"')
    script = f'''
    set theFile to choose file name with prompt "保存新 Q&A 文件" default name "{safe_name}"
    return POSIX path of theFile
    '''
    try:
        result = subprocess.run(
            ["osascript", "-e", script],
            capture_output=True, text=True, timeout=120,
        )
        if result.returncode == 0:
            path = result.stdout.strip()
            if path:
                return path
    except (FileNotFoundError, subprocess.TimeoutExpired):
        pass
    return None


def load_xlsx(path: str):
    """加载 xlsx,返回 (workbook, worksheet, header, data_rows)"""
    if not os.path.exists(path):
        raise FileNotFoundError(f"文件不存在: {path}")
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return wb, ws, [], []
    return wb, ws, list(rows[0]), [list(r) for r in rows[1:]]


def validate_xlsx(path: str) -> tuple[bool, str]:
    """校验 xlsx 是否符合 Q&A 格式
    返回 (是否有效, 错误信息);有效时错误信息为空字符串
    """
    if not os.path.exists(path):
        return False, f"文件不存在: {path}"
    if not path.lower().endswith(".xlsx"):
        return False, f"文件后缀不是 .xlsx(只支持 Excel 2007+ 格式,不支持 .xls/.csv)"
    try:
        # 不开 read_only,避免触发 openpyxl 对 extended metadata 的 lazy import
        # (在某些环境下会报 No module named 'openpyxl.packaging.extended')
        wb = load_workbook(path, data_only=True)
    except Exception as e:
        return False, f"无法打开 xlsx 文件(可能已损坏或被加密): {e}"

    try:
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
    finally:
        wb.close()

    if not rows:
        return False, "文件为空,没有任何行"

    # 表头校验
    header = [str(c or "").strip() for c in rows[0]]
    if len(header) < 3:
        return False, f"表头至少需要 3 列(序号 / 问题 / 答案),当前只有 {len(header)} 列:{header}"

    # 识别关键列(容错:中英文都支持)
    seq_aliases = {"序号", "编号", "id", "ID", "Id", "Seq", "seq", "No", "no", "NO", "#"}
    q_aliases = {"问题", "Question", "question", "Q", "q"}
    a_aliases = {"答案", "回答", "Answer", "answer", "A", "a"}

    seq_col = next((i for i, h in enumerate(header) if h in seq_aliases), None)
    q_col = next((i for i, h in enumerate(header) if h in q_aliases), None)
    a_col = next((i for i, h in enumerate(header) if h in a_aliases), None)

    if q_col is None:
        return False, f"表头找不到「问题」列(当前表头: {header})"
    if a_col is None:
        return False, f"表头找不到「答案」列(当前表头: {header})"
    if seq_col is None:
        return False, f"表头找不到「序号」列(当前表头: {header})"

    # 数据行校验
    data_rows = rows[1:]
    if not data_rows:
        return False, "文件只有表头,没有任何数据行"

    for i, r in enumerate(data_rows, start=2):
        if not r:
            return False, f"第 {i} 行为空"
        # 至少要能取到 q_col 和 a_col
        if max(q_col, a_col, seq_col) >= len(r):
            return False, f"第 {i} 行列数不足(只有 {len(r)} 列,需要至少 {max(q_col, a_col, seq_col)+1} 列)"
        q = r[q_col]
        if q is None or not str(q).strip():
            return False, f"第 {i} 行「问题」列为空"
        # 序号应该是数字(或可解析为数字)
        seq_val = r[seq_col]
        if seq_val is not None and not isinstance(seq_val, (int, float)):
            try:
                float(seq_val)
            except (ValueError, TypeError):
                return False, f"第 {i} 行「序号」不是数字: {seq_val!r}"
        # 答案可以为空(允许空答案)
    return True, ""


def find_qa_columns(header: list) -> tuple[int, int, int]:
    """从表头识别 序号/问题/答案 列的索引"""
    seq_aliases = {"序号", "编号", "id", "ID", "Id", "Seq", "seq", "No", "no", "NO", "#"}
    q_aliases = {"问题", "Question", "question", "Q", "q"}
    a_aliases = {"答案", "回答", "Answer", "answer", "A", "a"}
    seq_col = next((i for i, h in enumerate(header) if str(h or "").strip() in seq_aliases), 0)
    q_col = next((i for i, h in enumerate(header) if str(h or "").strip() in q_aliases), 1)
    a_col = next((i for i, h in enumerate(header) if str(h or "").strip() in a_aliases), 2)
    return seq_col, q_col, a_col


def list_all() -> list[dict]:
    """列出当前 xlsx 的所有 Q&A(自动识别列位置)"""
    path = XLSX_PATH["path"]
    if not path or not os.path.exists(path):
        return []
    _, _, header, data = load_xlsx(path)
    if not header:
        return []
    seq_col, q_col, a_col = find_qa_columns(header)
    out = []
    for r in data:
        # 缺的列填默认值
        def get(col, default=""):
            return r[col] if col < len(r) and r[col] is not None else default
        out.append({
            "seq": get(seq_col, 0),
            "question": get(q_col, ""),
            "answer": get(a_col, ""),
        })
    return out


def find_seq(question: str) -> int | None:
    """根据问题找已存在的序号"""
    for qa in list_all():
        if qa["question"] == question:
            return qa["seq"]
    return None


def write_xlsx(rows: list[list], header: list | None = None):
    """重写整个 xlsx(简单可靠,数据量小不影响)

    openpyxl 的 ws.delete_rows(1, max_row) 偶尔会留下表头残影,
    这里用更稳的方式:删 worksheet,重建。
    如果没传 header,从原文件读出来。
    """
    path = XLSX_PATH["path"]
    if not path:
        raise RuntimeError("xlsx 路径未设置")
    wb = load_workbook(path)
    old_ws = wb.active
    old_title = old_ws.title
    if header is None:
        # 尝试从原文件第一行读表头
        try:
            header = list(next(old_ws.iter_rows(min_row=1, max_row=1, values_only=True)))
        except StopIteration:
            header = ["序号", "问题", "答案"]
    wb.remove(old_ws)
    ws = wb.create_sheet(old_title)
    ws.append(header)
    for r in rows:
        ws.append(r)
    wb.save(path)


# ---------- 路由 ----------

@app.route("/")
def index():
    return render_template("index.html")


@app.route("/api/info")
def api_info():
    path = XLSX_PATH["path"]
    exists = path and os.path.exists(path)
    return jsonify({
        "path": path,
        "loaded": bool(exists),
        "answer_max": ANSWER_MAX,
    })


@app.route("/api/qa", methods=["GET"])
def api_list_qa():
    if not XLSX_PATH["path"]:
        return jsonify({"error": "xlsx 未加载"}), 400
    items = list_all()
    keyword = (request.args.get("q") or "").strip()
    if keyword:
        items = [qa for qa in items
                 if keyword in qa["question"] or keyword in qa["answer"]]
    return jsonify({"items": items, "total": len(items)})


@app.route("/api/qa", methods=["POST"])
def api_add_qa():
    if not XLSX_PATH["path"]:
        return jsonify({"error": "xlsx 未加载"}), 400

    data = request.get_json(silent=True) or {}
    question = (data.get("question") or "").strip()
    answer = data.get("answer") or ""

    if not question:
        return jsonify({"error": "问题不能为空"}), 400

    answer_len = count_codepoints(answer)
    if answer_len > ANSWER_MAX:
        return jsonify({
            "error": f"答案超过 {ANSWER_MAX} 字符限制(当前 {answer_len})",
            "answer_length": answer_len,
        }), 400

    # 检查重复
    existing = list_all()
    dup_seq = find_seq(question)
    if dup_seq is not None and not data.get("force_overwrite"):
        return jsonify({
            "duplicate": True,
            "existing_seq": dup_seq,
            "message": "问题已存在",
        }), 409

    # 写 xlsx
    if dup_seq is not None and data.get("force_overwrite"):
        # 覆盖: 找到那行,只改答案
        new_rows = [[qa["seq"], qa["question"], answer if qa["question"] == question else qa["answer"]]
                    for qa in existing]
        write_xlsx(new_rows)
        return jsonify({"ok": True, "action": "overwritten", "seq": dup_seq})

    # 新增: 接续最大序号
    max_seq = 0
    for qa in existing:
        s = qa["seq"]
        if isinstance(s, (int, float)):
            max_seq = max(max_seq, int(s))
    new_seq = max_seq + 1

    new_rows = [[qa["seq"], qa["question"], qa["answer"]] for qa in existing]
    new_rows.append([new_seq, question, answer])
    write_xlsx(new_rows)

    return jsonify({"ok": True, "action": "added", "seq": new_seq})


@app.route("/api/qa/<int:seq>", methods=["PUT"])
def api_update_qa(seq):
    """编辑指定序号的问题/答案(序号不变)"""
    if not XLSX_PATH["path"]:
        return jsonify({"error": "xlsx 未加载"}), 400

    data = request.get_json(silent=True) or {}
    new_question = (data.get("question") or "").strip()
    new_answer = data.get("answer") or ""

    if not new_question:
        return jsonify({"error": "问题不能为空"}), 400

    answer_len = count_codepoints(new_answer)
    if answer_len > ANSWER_MAX:
        return jsonify({
            "error": f"答案超过 {ANSWER_MAX} 字符限制(当前 {answer_len})",
            "answer_length": answer_len,
        }), 400

    qas = list_all()
    target = next((q for q in qas if q["seq"] == seq), None)
    if not target:
        return jsonify({"error": f"找不到序号 #{seq}"}), 404

    # 重复检测(排除自己)
    conflict = next((q for q in qas if q["seq"] != seq and q["question"] == new_question), None)
    if conflict:
        return jsonify({
            "duplicate": True,
            "conflict_seq": conflict["seq"],
            "message": f"问题与已存在的 #{conflict['seq']} 冲突",
        }), 409

    # 写回
    new_rows = []
    for qa in qas:
        if qa["seq"] == seq:
            new_rows.append([seq, new_question, new_answer])
        else:
            new_rows.append([qa["seq"], qa["question"], qa["answer"]])

    try:
        write_xlsx(new_rows)
    except Exception as e:
        return jsonify({"error": f"写入失败: {e}"}), 500

    return jsonify({"ok": True, "seq": seq})


@app.route("/api/xlsx", methods=["POST"])
def api_set_xlsx():
    data = request.get_json(silent=True) or {}
    path = (data.get("path") or "").strip()
    if not path:
        return jsonify({"error": "路径不能为空"}), 400
    path = os.path.expanduser(path)
    if not os.path.exists(path):
        return jsonify({"error": f"文件不存在: {path}"}), 404
    ok, err = validate_xlsx(path)
    if not ok:
        return jsonify({"error": f"文件格式校验失败:{err}"}), 400
    XLSX_PATH["path"] = os.path.abspath(path)
    return jsonify({"ok": True, "path": XLSX_PATH["path"]})


@app.route("/api/pick-file", methods=["POST"])
def api_pick_file():
    """后端弹 macOS 原生文件选择框,返回选中的绝对路径(只允许 .xlsx)"""
    if sys.platform != "darwin":
        return jsonify({"error": "此功能仅支持 macOS(其他平台请手动输入路径)"}), 400
    path = pick_file_macos()
    if not path:
        return jsonify({"cancelled": True, "path": None})
    if not path.lower().endswith(".xlsx"):
        return jsonify({"error": f"只支持 .xlsx 文件(你选的是: {os.path.basename(path)})"}), 400
    return jsonify({"ok": True, "path": path})


@app.route("/api/pick-save-path", methods=["POST"])
def api_pick_save_path():
    """后端弹 macOS 原生"保存"对话框,返回选中的路径(.xlsx)"""
    if sys.platform != "darwin":
        return jsonify({"error": "此功能仅支持 macOS(其他平台请手动输入路径)"}), 400
    data = request.get_json(silent=True) or {}
    default_name = (data.get("default_name") or "问答.xlsx").strip()
    if not default_name.lower().endswith(".xlsx"):
        default_name += ".xlsx"
    path = pick_save_path_macos(default_name)
    if not path:
        return jsonify({"cancelled": True, "path": None})
    if not path.lower().endswith(".xlsx"):
        path += ".xlsx"
    return jsonify({"ok": True, "path": path})


@app.route("/api/create", methods=["POST"])
def api_create():
    """创建新的 xlsx(只含表头),并切换到新文件"""
    data = request.get_json(silent=True) or {}
    path = (data.get("path") or "").strip()
    if not path:
        return jsonify({"error": "路径不能为空"}), 400
    path = os.path.expanduser(path)
    if not path.lower().endswith(".xlsx"):
        path += ".xlsx"

    # 目录不存在则创建
    parent = os.path.dirname(path)
    if parent and not os.path.exists(parent):
        try:
            os.makedirs(parent, exist_ok=True)
        except Exception as e:
            return jsonify({"error": f"无法创建目录 {parent}: {e}"}), 400

    # 文件已存在 → 拒绝(避免误覆盖)
    if os.path.exists(path):
        return jsonify({"error": f"文件已存在,不会覆盖: {path}"}), 400

    # 创建
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws.append(["序号", "问题", "答案"])
        # 设置表头样式(加粗)
        from openpyxl.styles import Font
        for cell in ws[1]:
            cell.font = Font(bold=True)
        # 列宽
        ws.column_dimensions["A"].width = 8
        ws.column_dimensions["B"].width = 40
        ws.column_dimensions["C"].width = 80
        wb.save(path)
    except Exception as e:
        return jsonify({"error": f"创建文件失败: {e}"}), 500

    XLSX_PATH["path"] = os.path.abspath(path)
    return jsonify({"ok": True, "path": XLSX_PATH["path"]})


@app.route("/api/merge", methods=["POST"])
def api_merge():
    """合并另一个 xlsx 到当前文件,合并后序号重制为 1..N"""
    if not XLSX_PATH["path"]:
        return jsonify({"error": "xlsx 未加载"}), 400

    data = request.get_json(silent=True) or {}
    source_path = (data.get("source_path") or "").strip()
    if not source_path:
        return jsonify({"error": "源文件路径不能为空"}), 400
    source_path = os.path.expanduser(source_path)
    if not os.path.exists(source_path):
        return jsonify({"error": f"源文件不存在: {source_path}"}), 404
    if os.path.abspath(source_path) == os.path.abspath(XLSX_PATH["path"]):
        return jsonify({"error": "源文件和目标文件相同,无需合并"}), 400
    ok, err = validate_xlsx(source_path)
    if not ok:
        return jsonify({"error": f"源文件格式校验失败:{err}"}), 400

    # 读源文件
    try:
        _, _, _, source_rows = load_xlsx(source_path)
    except Exception as e:
        return jsonify({"error": f"读取源文件失败: {e}"}), 400

    # 过滤空行
    source_qas = []
    for r in source_rows:
        if not r or len(r) < 3:
            continue
        q = (r[1] or "").strip()
        if not q:
            continue
        source_qas.append([q, r[2] or ""])

    # 读目标(当前)
    target_qas = list_all()
    target_questions = {(qa["question"] or "").strip() for qa in target_qas}

    # 合并:目标先,源追加(去重:问题相同的跳过)
    merged = [[qa["question"] or "", qa["answer"] or ""] for qa in target_qas]
    added = 0
    skipped = 0
    for q, a in source_qas:
        if q in target_questions:
            skipped += 1
            continue
        merged.append([q, a])
        target_questions.add(q)  # 防源文件自身有重复
        added += 1

    # 重新编号 1..N(写回时再赋序号)

    # 备份当前文件
    import shutil
    from datetime import datetime
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_path = XLSX_PATH["path"] + f".backup_{ts}"
    try:
        shutil.copy2(XLSX_PATH["path"], backup_path)
    except Exception as e:
        return jsonify({"error": f"备份失败: {e}"}), 500

    # 写回(重新编号)
    try:
        write_xlsx([[i + 1, q, a] for i, (q, a) in enumerate(merged)])
    except Exception as e:
        return jsonify({"error": f"写入失败: {e},已保留备份 {backup_path}"}), 500

    return jsonify({
        "ok": True,
        "backup_path": backup_path,
        "source_total": len(source_qas),
        "target_before": len(target_qas),
        "added": added,
        "skipped": skipped,
        "total_after": len(merged),
    })


@app.route("/api/preview-merge", methods=["POST"])
def api_preview_merge():
    """预览合并:不实际合并,只返回源/目标的统计,供前端展示"""
    if not XLSX_PATH["path"]:
        return jsonify({"error": "xlsx 未加载"}), 400
    data = request.get_json(silent=True) or {}
    source_path = (data.get("source_path") or "").strip()
    if not source_path:
        return jsonify({"error": "源文件路径不能为空"}), 400
    source_path = os.path.expanduser(source_path)
    if not os.path.exists(source_path):
        return jsonify({"error": f"源文件不存在: {source_path}"}), 404
    if os.path.abspath(source_path) == os.path.abspath(XLSX_PATH["path"]):
        return jsonify({"error": "源文件和目标文件相同"}), 400
    ok, err = validate_xlsx(source_path)
    if not ok:
        return jsonify({"error": f"源文件格式校验失败:{err}"}), 400

    try:
        _, _, _, source_rows = load_xlsx(source_path)
    except Exception as e:
        return jsonify({"error": f"读取源文件失败: {e}"}), 400

    source_qas = []
    for r in source_rows:
        if not r or len(r) < 3:
            continue
        q = (r[1] or "").strip()
        if not q:
            continue
        source_qas.append(q)

    target_qas = list_all()
    target_questions = {(qa["question"] or "").strip() for qa in target_qas}
    overlap = sum(1 for q in source_qas if q in target_questions)

    return jsonify({
        "source_total": len(source_qas),
        "target_total": len(target_qas),
        "will_add": len(source_qas) - overlap,
        "will_skip": overlap,
        "merged_total": len(target_qas) + (len(source_qas) - overlap),
    })


# ---------- 启动 ----------

def main():
    if len(sys.argv) > 1:
        path = os.path.expanduser(sys.argv[1])
    elif sys.platform == "darwin":
        print("未指定 xlsx 路径,弹出文件选择框...")
        path = pick_file_macos()
        if not path:
            print("未选择文件,退出。")
            print("提示: 也可以这样指定路径启动: python qa_web.py /path/to/file.xlsx")
            sys.exit(1)
    else:
        print("请指定 xlsx 路径: python qa_web.py /path/to/file.xlsx")
        sys.exit(1)

    if not os.path.exists(path):
        print(f"文件不存在: {path}")
        sys.exit(1)

    ok, err = validate_xlsx(path)
    if not ok:
        print(f"✗ 文件格式校验失败: {err}")
        sys.exit(1)

    XLSX_PATH["path"] = os.path.abspath(path)
    print(f"✓ 已加载: {XLSX_PATH['path']}")
    print(f"✓ 浏览器打开: http://localhost:5000")
    print(f"  (按 Ctrl+C 停止)")
    app.run(host="127.0.0.1", port=5000, debug=False)


if __name__ == "__main__":
    main()
